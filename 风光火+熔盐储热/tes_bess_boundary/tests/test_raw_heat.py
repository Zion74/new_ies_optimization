from __future__ import annotations

import hashlib
from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook

from tes_bess_boundary.raw_heat import audit_heat_workbooks, parse_heat_workbook


DATA_ROOT = Path(__file__).resolve().parents[2]


def _write_heat_workbook(path: Path, timestamps: list[datetime]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    for offset, timestamp in enumerate(timestamps, start=9):
        sheet.cell(
            row=offset,
            column=1,
            value=(
                f"{timestamp:%Y-%m-%d} {timestamp:%H:%M:%S} "
                "360 36 18 1 2"
            ),
        )
    sheet.cell(row=9, column=10, value="auxiliary statistics must be ignored")
    workbook.save(path)


def test_parser_reads_only_column_a_even_when_auxiliary_cells_are_populated(
    tmp_path: Path,
) -> None:
    path = tmp_path / "heat.xlsx"
    _write_heat_workbook(path, [datetime(2024, 1, 1, 0, 0)])

    parsed = parse_heat_workbook(path)

    assert len(parsed.records) == 1
    assert parsed.records[0].timestamp == datetime(2024, 1, 1, 0, 0)
    assert parsed.records[0].resident_gj_per_h == pytest.approx(360.0)


def test_parsed_record_preserves_source_provenance(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    _write_heat_workbook(path, [datetime(2024, 1, 1, 0, 0)])

    record = parse_heat_workbook(path).records[0]

    assert record.source_path == path
    assert record.source_sha256 == hashlib.sha256(path.read_bytes()).hexdigest()
    assert record.source_sheet == "Sheet"
    assert record.source_row == 9
    assert record.source_raw_cell == "2024-01-01 00:00:00 360 36 18 1 2"


def test_non_ten_minute_boundary_point_is_excluded(tmp_path: Path) -> None:
    path = tmp_path / "boundary.xlsx"
    timestamps = [
        datetime(2024, 1, 1, 23, 0) + timedelta(minutes=10 * index)
        for index in range(6)
    ]
    timestamps.append(datetime(2024, 1, 1, 23, 59, 59))
    _write_heat_workbook(path, timestamps)

    audit = audit_heat_workbooks([path], year=2024, require_complete_year=False)

    assert audit.grid_record_count == 6
    assert audit.excluded_non_grid_count == 1
    assert audit.hourly_sample_counts[datetime(2024, 1, 1, 23, 0)] == 6


def test_dated_parse_error_preserves_workbook_row_and_raw_value(
    tmp_path: Path,
) -> None:
    path = tmp_path / "malformed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=9, column=1, value="2024-01-01 00:00:00 bad payload")
    workbook.save(path)

    parsed = parse_heat_workbook(path)

    assert parsed.parse_error_count == 1
    assert parsed.issues[0].row_number == 9
    assert parsed.issues[0].raw_value == "2024-01-01 00:00:00 bad payload"
    assert "five values" in parsed.issues[0].reason


def test_invalid_date_like_row_is_reported_instead_of_treated_as_a_header(
    tmp_path: Path,
) -> None:
    path = tmp_path / "invalid_timestamp.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(
        row=9,
        column=1,
        value="2024-02-30 00:00:00 1 2 3 4 5",
    )
    workbook.save(path)

    parsed = parse_heat_workbook(path)

    assert parsed.records == ()
    assert parsed.parse_error_count == 1
    assert parsed.issues[0].row_number == 9
    assert "timestamp" in parsed.issues[0].reason


def test_clean_partial_audit_requires_explicit_quality_review(
    tmp_path: Path,
) -> None:
    path = tmp_path / "clean_hour.xlsx"
    _write_heat_workbook(
        path,
        [
            datetime(2024, 1, 1, 0, 0) + timedelta(minutes=10 * index)
            for index in range(6)
        ],
    )

    unconfirmed = audit_heat_workbooks(
        [path],
        year=2024,
        require_complete_year=False,
    )
    confirmed = audit_heat_workbooks(
        [path],
        year=2024,
        require_complete_year=False,
        quality_review_confirmed=True,
    )

    assert not unconfirmed.formal_ready
    assert confirmed.formal_ready
    assert confirmed.zero_heat_sample_count == 0
    assert confirmed.longest_zero_heat_run_samples == 0


def test_legacy_zero_signal_audit_requires_all_five_measurements_to_be_zero(
    tmp_path: Path,
) -> None:
    path = tmp_path / "zero_definition.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(
        row=9,
        column=1,
        value="2024-01-01 00:00:00 10 -10 0 1 -1",
    )
    sheet.cell(
        row=10,
        column=1,
        value="2024-01-01 00:10:00 0 0 0 0 0",
    )
    workbook.save(path)

    audit = audit_heat_workbooks(
        [path],
        year=2024,
        require_complete_year=False,
    )

    assert audit.all_signal_zero_sample_count == 1
    assert audit.longest_all_signal_zero_run_samples == 1
    assert audit.zero_heat_sample_count == 1
    assert audit.longest_zero_heat_run_samples == 1


@pytest.mark.data_integration
def test_real_raw_heat_data_has_a_complete_strict_ten_minute_grid() -> None:
    base = DATA_ROOT / "杨凌机组数据" / "杨凌机组数据" / "热功率"
    paths = [
        base / "2024.1.1-2024.6.30杨凌供热数据.xlsx",
        base / "2024.7.1-2024.12.31杨凌供热数据.xlsx",
    ]

    audit = audit_heat_workbooks(paths, year=2024)

    assert audit.grid_record_count == 52704
    assert audit.identical_duplicate_count == 1
    assert audit.excluded_non_grid_count == 2
    assert audit.sample_count_distribution == {6: 8784}
    assert audit.negative_counts["resident_gj_per_h"] == 2050
    assert audit.negative_counts["dongfang_gj_per_h"] == 29
    assert audit.sentinel_counts["oldcity_gj_per_h"] == 1
    assert not audit.formal_ready
