"""Source-level audit for Yangling 10-minute heat workbooks.

Only column A is parsed. Auxiliary statistics elsewhere on the worksheet are
deliberately ignored because they corrupted the legacy row-token parser.
"""

from __future__ import annotations

import hashlib
import re
from collections import Counter
from dataclasses import dataclass, replace
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook


HEAT_VALUE_FIELDS = (
    "resident_gj_per_h",
    "dongfang_gj_per_h",
    "oldcity_gj_per_h",
    "dongfang_flow",
    "oldcity_flow",
)

_DATE_LIKE_TOKEN = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$")
_TIME_LIKE_TOKEN = re.compile(r"^\d{1,2}:\d{2}(?::\d{2})?$")


@dataclass(frozen=True)
class HeatRecord:
    timestamp: datetime
    resident_gj_per_h: float
    dongfang_gj_per_h: float
    oldcity_gj_per_h: float
    dongfang_flow: float
    oldcity_flow: float
    source_path: Path | None = None
    source_sha256: str | None = None
    source_sheet: str | None = None
    source_row: int | None = None
    source_raw_cell: str | None = None

    @property
    def values(self) -> tuple[float, ...]:
        return tuple(float(getattr(self, field)) for field in HEAT_VALUE_FIELDS)


@dataclass(frozen=True)
class HeatParseIssue:
    path: Path
    row_number: int
    raw_value: str
    reason: str


@dataclass(frozen=True)
class ParsedHeatWorkbook:
    path: Path
    records: tuple[HeatRecord, ...]
    issues: tuple[HeatParseIssue, ...]

    @property
    def parse_error_count(self) -> int:
        return len(self.issues)


@dataclass(frozen=True)
class RawHeatAudit:
    year: int
    grid_record_count: int
    expected_grid_record_count: int
    identical_duplicate_count: int
    excluded_non_grid_count: int
    parse_error_count: int
    parse_issues: tuple[HeatParseIssue, ...]
    missing_grid_count: int
    hourly_sample_counts: Mapping[datetime, int]
    sample_count_distribution: Mapping[int, int]
    negative_counts: Mapping[str, int]
    sentinel_counts: Mapping[str, int]
    all_signal_zero_sample_count: int
    longest_all_signal_zero_run_samples: int
    quality_review_confirmed: bool
    formal_ready: bool

    @property
    def zero_heat_sample_count(self) -> int:
        """Backward-compatible alias for all-five-signal zero samples."""

        return self.all_signal_zero_sample_count

    @property
    def longest_zero_heat_run_samples(self) -> int:
        """Backward-compatible alias for the all-five-signal zero run."""

        return self.longest_all_signal_zero_run_samples


class ConflictingDuplicateError(ValueError):
    """Raised when one timestamp has different source values."""


def _parse_timestamp(date_token: str, time_token: str) -> datetime:
    normalized_date = date_token.strip().replace("/", "-")
    return datetime.fromisoformat(f"{normalized_date} {time_token.strip()}")


def _parse_column_a_cell(raw: object) -> HeatRecord | None:
    if raw is None:
        return None
    tokens = str(raw).replace("\t", " ").split()
    if len(tokens) < 2:
        return None
    try:
        timestamp = _parse_timestamp(tokens[0], tokens[1])
    except ValueError as exc:
        if _DATE_LIKE_TOKEN.fullmatch(tokens[0]) and _TIME_LIKE_TOKEN.fullmatch(
            tokens[1]
        ):
            raise ValueError("heat-data timestamp is invalid") from exc
        return None
    if len(tokens) != 7:
        raise ValueError("dated heat-data row must contain date, time, and five values")
    try:
        values = tuple(float(token) for token in tokens[2:])
    except ValueError as exc:
        raise ValueError("heat-data values must be numeric") from exc
    return HeatRecord(timestamp, *values)


def parse_heat_workbook(path: str | Path) -> ParsedHeatWorkbook:
    """Parse timestamp and five measurements from column A of the first sheet."""

    workbook_path = Path(path)
    source_sha256 = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    records: list[HeatRecord] = []
    issues: list[HeatParseIssue] = []
    try:
        for row_number, (raw,) in enumerate(
            sheet.iter_rows(min_col=1, max_col=1, values_only=True),
            start=1,
        ):
            try:
                record = _parse_column_a_cell(raw)
            except ValueError as exc:
                issues.append(
                    HeatParseIssue(
                        path=workbook_path,
                        row_number=row_number,
                        raw_value=str(raw),
                        reason=str(exc),
                    )
                )
                continue
            if record is not None:
                records.append(
                    replace(
                        record,
                        source_path=workbook_path,
                        source_sha256=source_sha256,
                        source_sheet=sheet.title,
                        source_row=row_number,
                        source_raw_cell=str(raw),
                    )
                )
    finally:
        workbook.close()
    return ParsedHeatWorkbook(workbook_path, tuple(records), tuple(issues))


def _is_ten_minute_grid(timestamp: datetime) -> bool:
    return (
        timestamp.minute % 10 == 0
        and timestamp.second == 0
        and timestamp.microsecond == 0
    )


def _hourly_index(year: int) -> tuple[datetime, ...]:
    start = datetime(year, 1, 1)
    count = int((datetime(year + 1, 1, 1) - start).total_seconds() // 3600)
    return tuple(start + timedelta(hours=index) for index in range(count))


def _ten_minute_index(year: int) -> set[datetime]:
    start = datetime(year, 1, 1)
    count = len(_hourly_index(year)) * 6
    return {start + timedelta(minutes=10 * index) for index in range(count)}


def audit_heat_workbooks(
    paths: Sequence[str | Path],
    *,
    year: int,
    require_complete_year: bool = True,
    sentinel_abs_threshold: float = 1_000_000.0,
    quality_review_confirmed: bool = False,
) -> RawHeatAudit:
    """Audit strict grid coverage without silently repairing physical values."""

    parsed = [parse_heat_workbook(path) for path in paths]
    parse_error_count = sum(workbook.parse_error_count for workbook in parsed)
    parse_issues = tuple(
        issue for workbook in parsed for issue in workbook.issues
    )
    start = datetime(year, 1, 1)
    end = datetime(year + 1, 1, 1)
    excluded_non_grid_count = 0
    identical_duplicate_count = 0
    merged: dict[datetime, HeatRecord] = {}

    for workbook in parsed:
        for record in workbook.records:
            if not start <= record.timestamp < end:
                continue
            if not _is_ten_minute_grid(record.timestamp):
                excluded_non_grid_count += 1
                continue
            existing = merged.get(record.timestamp)
            if existing is None:
                merged[record.timestamp] = record
            elif existing.values == record.values:
                identical_duplicate_count += 1
            else:
                raise ConflictingDuplicateError(
                    f"Conflicting heat records at {record.timestamp.isoformat()}"
                )

    expected_grid = _ten_minute_index(year)
    missing_grid_count = len(expected_grid.difference(merged))
    hourly_counter = Counter(
        timestamp.replace(minute=0, second=0, microsecond=0) for timestamp in merged
    )
    if require_complete_year:
        hourly_sample_counts = {
            hour: hourly_counter.get(hour, 0) for hour in _hourly_index(year)
        }
    else:
        hourly_sample_counts = dict(sorted(hourly_counter.items()))
    sample_count_distribution = dict(
        sorted(Counter(hourly_sample_counts.values()).items())
    )

    negative_counts = {
        field: sum(getattr(record, field) < 0 for record in merged.values())
        for field in HEAT_VALUE_FIELDS
    }
    sentinel_counts = {
        field: sum(
            abs(getattr(record, field)) >= sentinel_abs_threshold
            for record in merged.values()
        )
        for field in HEAT_VALUE_FIELDS
    }
    all_signal_zero_timestamps = sorted(
        timestamp
        for timestamp, record in merged.items()
        if all(value == 0.0 for value in record.values)
    )
    longest_zero_run = 0
    current_zero_run = 0
    previous_zero_timestamp: datetime | None = None
    for timestamp in all_signal_zero_timestamps:
        if (
            previous_zero_timestamp is not None
            and timestamp - previous_zero_timestamp == timedelta(minutes=10)
        ):
            current_zero_run += 1
        else:
            current_zero_run = 1
        longest_zero_run = max(longest_zero_run, current_zero_run)
        previous_zero_timestamp = timestamp

    if require_complete_year:
        complete = (
            missing_grid_count == 0
            and len(merged) == len(expected_grid)
            and all(count == 6 for count in hourly_sample_counts.values())
        )
    else:
        complete = bool(merged) and all(
            count == 6 for count in hourly_sample_counts.values()
        )
    formal_ready = (
        complete
        and parse_error_count == 0
        and not any(negative_counts.values())
        and not any(sentinel_counts.values())
        and quality_review_confirmed
    )

    return RawHeatAudit(
        year=year,
        grid_record_count=len(merged),
        expected_grid_record_count=len(expected_grid),
        identical_duplicate_count=identical_duplicate_count,
        excluded_non_grid_count=excluded_non_grid_count,
        parse_error_count=parse_error_count,
        parse_issues=parse_issues,
        missing_grid_count=missing_grid_count,
        hourly_sample_counts=hourly_sample_counts,
        sample_count_distribution=sample_count_distribution,
        negative_counts=negative_counts,
        sentinel_counts=sentinel_counts,
        all_signal_zero_sample_count=len(all_signal_zero_timestamps),
        longest_all_signal_zero_run_samples=longest_zero_run,
        quality_review_confirmed=quality_review_confirmed,
        formal_ready=formal_ready,
    )
