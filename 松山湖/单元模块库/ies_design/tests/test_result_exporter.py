import csv
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from defaults_resolver import DefaultsResolver
from schema_validator import ValidationResult
from result_exporter import ResultExporter
from scenario_loader import ScenarioLoader


def write_pareto(path: Path):
    path.parent.mkdir(parents=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Solution_ID", "Economic_Cost", "Matching_Index", "PV", "GT", "EC"])
        writer.writerow([0, 100.0, 30.0, 10.0, 20.0, 30.0])
        writer.writerow([1, 80.0, 50.0, 11.0, 21.0, 31.0])
        writer.writerow([2, 999999999.0, 1.0, 12.0, 22.0, 32.0])


def test_exports_standard_design_files_and_filters_infeasible_recommendations():
    with tempfile.TemporaryDirectory() as tmp:
        result_dir = Path(tmp)
        write_pareto(result_dir / "Euclidean" / "Pareto_Euclidean.csv")
        scenario = ScenarioLoader.load(ROOT / "scenarios" / "songshan_lake" / "scenario.yaml")
        resolved = DefaultsResolver(ROOT / "defaults").resolve(scenario)

        validation = ValidationResult(warnings=["demo warning"])

        outputs = ResultExporter.export(result_dir, resolved, validation=validation)

        assert outputs["pareto_solutions"].exists()
        assert outputs["design_summary"].exists()
        assert outputs["design_summary_wide"].exists()
        assert outputs["design_summary_xlsx"].exists()
        assert outputs["design_report"].exists()
        assert outputs["resolved_scenario"].exists()
        assert outputs["validation_report"].exists()

        wide = outputs["design_summary_wide"].read_text(encoding="utf-8")
        assert "min_cost" in wide
        assert "min_matching" in wide
        assert "999999999" not in wide
        assert "PV" in wide

        long_table = outputs["design_summary"].read_text(encoding="utf-8")
        assert "device_capacity" in long_table
        assert "economic_cost" in long_table

        resolved_json = outputs["resolved_scenario"].read_text(encoding="utf-8")
        assert '"id": "songshan_lake"' in resolved_json

        validation_report = outputs["validation_report"].read_text(encoding="utf-8")
        assert "demo warning" in validation_report

        report = outputs["design_report"].read_text(encoding="utf-8")
        assert "## 输入数据" in report
        assert "## 系统结构" in report
        assert "## 优化设置" in report

        import openpyxl

        wb = openpyxl.load_workbook(outputs["design_summary_xlsx"], data_only=True)
        assert "summary_wide" in wb.sheetnames
        assert "summary_long" in wb.sheetnames
        assert "device_metadata" in wb.sheetnames
        metadata = wb["device_metadata"]
        headers = [cell.value for cell in metadata[1]]
        assert headers == [
            "solution_label",
            "device_id",
            "device_name",
            "device_type",
            "input_carriers",
            "output_carriers",
            "capacity_value",
            "unit",
            "is_default_device",
            "is_user_configured",
        ]
        rows = list(metadata.iter_rows(min_row=2, values_only=True))
        assert any(row[1] == "pv" and row[2] == "标准光伏" for row in rows)
        assert any(row[1] == "chp" and row[6] == 20 for row in rows)


if __name__ == "__main__":
    for name, fn in sorted(globals().items()):
        if name.startswith("test_"):
            fn()
            print(f"PASS {name}")
