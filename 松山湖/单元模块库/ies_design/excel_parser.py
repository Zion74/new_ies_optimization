from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class ExcelParseResult:
    scenario: dict[str, Any]
    warnings: list[str] = field(default_factory=list)
    tables: dict[str, list[dict[str, Any]]] = field(default_factory=dict)

    def export(self, output_dir: str | Path) -> dict[str, Path]:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        scenario_yaml = output_dir / "scenario.yaml"
        typical_profiles = output_dir / "typical_profiles.csv"
        input_resource_profiles = output_dir / "input_resource_profiles.csv"
        data_gaps = output_dir / "data_gaps.csv"
        warnings_file = output_dir / "excel_parse_warnings.txt"

        scenario_yaml.write_text(_dump_yaml(self.scenario), encoding="utf-8")
        _write_csv(typical_profiles, self.tables.get("load_profiles", []))
        _write_csv(input_resource_profiles, self.tables.get("input_resource_profiles", []))
        _write_csv(data_gaps, self.tables.get("data_gaps", []))
        warnings_file.write_text("\n".join(self.warnings) + ("\n" if self.warnings else ""), encoding="utf-8")

        return {
            "scenario_yaml": scenario_yaml,
            "typical_profiles": typical_profiles,
            "input_resource_profiles": input_resource_profiles,
            "data_gaps": data_gaps,
            "warnings": warnings_file,
        }


class ExcelScenarioParser:
    """Parse the collaborator-facing scenario collection workbook."""

    REQUIRED_SHEETS = [
        "01_场景信息",
        "02_场景能源配置",
        "03_月度用户需求",
        "04_用户负荷需求曲线",
        "05_能源输入与资源曲线",
        "06_候选设备配置",
        "07_价格与排放参数",
        "08_资料来源与缺口",
    ]

    @classmethod
    def parse(cls, workbook_path: str | Path) -> ExcelParseResult:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - exercised in environments without openpyxl
            raise RuntimeError("Excel parsing requires openpyxl. Use `uv run python ...` in this project.") from exc

        workbook_path = Path(workbook_path)
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        warnings: list[str] = []
        for sheet in cls.REQUIRED_SHEETS:
            if sheet not in wb.sheetnames:
                warnings.append(f"missing required sheet: {sheet}")

        scenario_info = _read_key_value_sheet(wb["01_场景信息"]) if "01_场景信息" in wb.sheetnames else {}
        energy_rows = _read_table(wb["02_场景能源配置"]) if "02_场景能源配置" in wb.sheetnames else []
        load_profiles = _read_table(wb["04_用户负荷需求曲线"]) if "04_用户负荷需求曲线" in wb.sheetnames else []
        input_profiles = _read_table(wb["05_能源输入与资源曲线"]) if "05_能源输入与资源曲线" in wb.sheetnames else []
        device_rows = _read_table(wb["06_候选设备配置"]) if "06_候选设备配置" in wb.sheetnames else []
        price_rows = _read_table(wb["07_价格与排放参数"]) if "07_价格与排放参数" in wb.sheetnames else []
        data_gaps = _read_table(wb["08_资料来源与缺口"]) if "08_资料来源与缺口" in wb.sheetnames else []

        scenario = _build_scenario(scenario_info, energy_rows, device_rows, price_rows, workbook_path, warnings)
        _warn_missing_required_info(scenario_info, warnings)
        _warn_empty_required_tables(load_profiles, input_profiles, warnings)

        return ExcelParseResult(
            scenario=scenario,
            warnings=warnings,
            tables={
                "load_profiles": load_profiles,
                "input_resource_profiles": input_profiles,
                "data_gaps": data_gaps,
            },
        )


def _read_key_value_sheet(ws: Any) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    data: dict[str, Any] = {}
    for row in rows[1:]:
        key = _clean(row[0] if len(row) > 0 else None)
        if not key:
            continue
        data[_canonical_field(str(key))] = _clean(row[1] if len(row) > 1 else None)
    return data


def _read_table(ws: Any) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_canonical_field(str(_clean(cell) or "")) for cell in rows[0]]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[idx]: _clean(row[idx] if idx < len(row) else None) for idx in range(len(headers)) if headers[idx]}
        if any(value not in (None, "") for value in item.values()):
            output.append(item)
    return output


def _build_scenario(
    info: dict[str, Any],
    energy_rows: list[dict[str, Any]],
    device_rows: list[dict[str, Any]],
    price_rows: list[dict[str, Any]],
    workbook_path: Path,
    warnings: list[str],
) -> dict[str, Any]:
    scenario_id = str(info.get("scenario_id") or "excel_scenario")
    scenario_name = str(info.get("scenario_name") or "Excel导入场景")
    scenario_type = str(info.get("scenario_type") or "industrial_park_pv_geothermal_waste_heat")

    carriers = _energy_carriers(energy_rows)
    devices = _devices(device_rows, warnings)
    prices = _prices(price_rows)

    if not devices:
        warnings.append("no enabled devices found in 06_候选设备配置; default current CCHP devices are not auto-enabled for Excel import")

    return {
        "schema_version": "1.0",
        "scenario": {
            "id": scenario_id,
            "name": scenario_name,
            "scenario_type": scenario_type,
            "description": f"由 Excel 模板导入: {workbook_path.name}",
            "location": info.get("location") or "",
            "currency": info.get("currency") or "CNY",
        },
        "system": {"template": _template_for_scenario_type(scenario_type)},
        "energy_carriers": carriers,
        "data": {
            "input_type": "excel_template",
            "source_workbook": str(workbook_path),
            "load_file": "",
            "typical_day_file": "",
        },
        "simulation": {
            "mode": "typical_days",
            "annualization": "weighted",
            "time_step_hours": 1,
            "dispatch_horizon_hours": 24,
        },
        "typical_day": {
            "source": "monthly_template",
            "file": "",
        },
        "prices": prices,
        "devices": devices,
        "optimization": {
            "mode": "test",
            "methods": ["euclidean"],
        },
    }


def _template_for_scenario_type(scenario_type: str) -> str:
    if scenario_type == "tobacco_factory_multi_energy":
        return "tobacco_factory_multi_energy"
    return "cchp_ehc_base"


def _energy_carriers(rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    groups = {"demand": "demands", "input": "inputs", "resource": "resources"}
    carriers = {"demands": [], "inputs": [], "resources": []}
    for row in rows:
        group_key = groups.get(str(row.get("item_group") or ""))
        item_id = row.get("item_id")
        if not group_key or not item_id:
            continue
        if _is_enabled(row.get("enabled"), row.get("required")):
            carriers[group_key].append(str(item_id))
    if not carriers["demands"]:
        carriers["demands"].append("electricity")
    return carriers


def _devices(rows: list[dict[str, Any]], warnings: list[str]) -> dict[str, dict[str, Any]]:
    devices: dict[str, dict[str, Any]] = {}
    for idx, row in enumerate(rows, start=2):
        device_id = row.get("device_id")
        library_id = row.get("library_id")
        if _is_enabled(row.get("enabled"), None) and (not device_id or not library_id):
            warnings.append(f"06_候选设备配置 row {idx} is enabled but missing device_id or library_id")
        if not device_id or not library_id or not _is_enabled(row.get("enabled"), None):
            continue
        config: dict[str, Any] = {
            "library_id": str(library_id),
            "enabled": True,
            "optimize_capacity": _to_bool(row.get("optimize_capacity"), default=True),
        }
        if row.get("capacity_ub") not in (None, ""):
            config["capacity_ub_kw"] = row.get("capacity_ub")
        if row.get("power_ub") not in (None, ""):
            config["power_ub_kw"] = row.get("power_ub")
        devices[str(device_id)] = config
    return devices


def _prices(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    prices: dict[str, dict[str, Any]] = {
        "electricity": {"type": "flat", "unit": "CNY_per_kWh", "value": None},
        "gas": {"type": "flat", "unit": "CNY_per_kWh", "value": None},
        "capacity_charge": {"unit": "CNY_per_kW_year", "value": 0},
    }
    for row in rows:
        param_type = row.get("param_type")
        param_id = row.get("param_id")
        if param_type == "price" and row.get("carrier_id") == "electricity":
            prices["electricity"] = {"type": "flat", "unit": row.get("unit") or "CNY_per_kWh", "value": row.get("value")}
        elif param_type == "price" and row.get("carrier_id") == "natural_gas":
            prices["gas"] = {"type": "flat", "unit": row.get("unit") or "CNY_per_kWh", "value": row.get("value")}
        elif param_type == "capacity_charge" or param_id == "capacity_charge":
            prices["capacity_charge"] = {"unit": row.get("unit") or "CNY_per_kW_year", "value": row.get("value") or 0}
    return prices


FIELD_ALIASES = {
    "场景编号": "scenario_id",
    "场景ID": "scenario_id",
    "场景id": "scenario_id",
    "编号": "scenario_id",
    "场景名称": "scenario_name",
    "名称": "scenario_name",
    "场景类型": "scenario_type",
    "类型": "scenario_type",
    "地点": "location",
    "位置": "location",
    "币种": "currency",
    "货币": "currency",
    "条目ID": "item_id",
    "条目编号": "item_id",
    "能源ID": "item_id",
    "能源载体": "carrier_id",
    "载体ID": "carrier_id",
    "分组": "item_group",
    "类型分组": "item_group",
    "项目分组": "item_group",
    "启用": "enabled",
    "是否启用": "enabled",
    "必需": "required",
    "是否必需": "required",
    "设备编号": "device_id",
    "设备ID": "device_id",
    "设备id": "device_id",
    "设备库编号": "library_id",
    "设备库ID": "library_id",
    "库设备ID": "library_id",
    "优化容量": "optimize_capacity",
    "是否优化容量": "optimize_capacity",
    "容量上限": "capacity_ub",
    "装机上限": "capacity_ub",
    "功率上限": "power_ub",
    "参数编号": "param_id",
    "参数ID": "param_id",
    "参数类型": "param_type",
    "数值": "value",
    "值": "value",
    "单位": "unit",
    "需求ID": "demand_id",
    "负荷ID": "demand_id",
    "输入ID": "input_id",
    "资源ID": "input_id",
}


def _canonical_field(value: str) -> str:
    text = value.strip()
    if not text:
        return text
    if text in FIELD_ALIASES:
        return FIELD_ALIASES[text]
    normalized = text.replace(" ", "_")
    return FIELD_ALIASES.get(normalized, normalized)


def _warn_missing_required_info(info: dict[str, Any], warnings: list[str]) -> None:
    for key in ["scenario_id", "scenario_name", "scenario_type"]:
        if not info.get(key):
            warnings.append(f"{key} is empty in 01_场景信息; using a placeholder value")


def _warn_empty_required_tables(
    load_profiles: list[dict[str, Any]],
    input_profiles: list[dict[str, Any]],
    warnings: list[str],
) -> None:
    if not any(row.get("demand_id") and row.get("value") not in (None, "") for row in load_profiles):
        warnings.append("04_用户负荷需求曲线 has no usable load values")
    if not any(row.get("input_id") and row.get("value") not in (None, "") for row in input_profiles):
        warnings.append("05_能源输入与资源曲线 has no usable input/resource values")


def _is_enabled(enabled: Any, required: Any) -> bool:
    if _to_bool(enabled, default=None) is not None:
        return bool(_to_bool(enabled, default=False))
    return _to_bool(required, default=False)


def _to_bool(value: Any, default: bool | None = False) -> bool | None:
    if value is None or value == "":
        return default
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "是", "启用"}:
        return True
    if text in {"false", "no", "n", "0", "否", "停用"}:
        return False
    return default


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    columns = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        if not columns:
            f.write("")
            return
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _dump_yaml(data: Any, indent: int = 0) -> str:
    lines: list[str] = []
    _dump_yaml_into(data, lines, indent)
    return "\n".join(lines) + "\n"


def _dump_yaml_into(data: Any, lines: list[str], indent: int) -> None:
    prefix = " " * indent
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, dict):
                lines.append(f"{prefix}{key}:")
                _dump_yaml_into(value, lines, indent + 2)
            elif isinstance(value, list):
                lines.append(f"{prefix}{key}: [{', '.join(_format_scalar(item) for item in value)}]")
            else:
                lines.append(f"{prefix}{key}: {_format_scalar(value)}")
    else:
        lines.append(f"{prefix}{_format_scalar(data)}")


def _format_scalar(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace('"', '\\"')
    return f'"{text}"'
