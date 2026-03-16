# 本程序编写时间：2023.02
# 用于求解热力网络稳态模型
# 流量单位：m3/h；压力单位：mH2O
# 本程序考虑了工频泵和变频泵特性曲线（利用二次函数拟合）
# 本程序考虑了电动阀门阻力-开度特性曲线（利用【R^(2(1-x))/K^2】形式特性曲线）
# 本程序用于求解多个断面下的水力工况

import numpy as np
import openpyxl
import math
import IHES_ReadNetworkPara_V2

# 定义求解水力模型的函数。可直接求解多个时间断面的水力工况
def Hydraulic_Calculation(CaseData, ResultFile, NumOfTime):

    #####    首先处理与调度时刻无关的计算过程
    RowOfNode = CaseData['Node'].shape[0]
    RowOfBranch = CaseData['Branch'].shape[0]

    # 生成增广节点-支路关联矩阵
    A = np.zeros(shape=(RowOfNode, RowOfBranch))
    for i in range(RowOfBranch):
        A[int(CaseData['Branch'][i][1]) - 1][i] = -1     # 从节点流出，则赋值为-1
        A[int(CaseData['Branch'][i][2]) - 1][i] = 1    # 流向节点，则赋值为1

    # 生成节点-支路关联矩阵,即将增广矩阵中平衡节点对应的行删除
    BalanceNodeIndex = []   # 存储平衡节点的行号
    LoadNodeIndex = []     # 存储载荷节点的行号
    for i in range(RowOfNode):
        if CaseData['Node'][i][1] == 0:
            BalanceNodeIndex.append(i)
        else:
            LoadNodeIndex.append(i)
    BalanceNodeIndex = np.array(BalanceNodeIndex)
    LoadNodeIndex = np.array(LoadNodeIndex)
    RowOfBalanceNode = BalanceNodeIndex.shape[0]
    RowOfLoadNode = LoadNodeIndex.shape[0]
    for i in range(RowOfBalanceNode):
        A = np.delete(A, BalanceNodeIndex[i]-i, axis = 0)   #此处的“-i”是因为每一删除一行后，A中总行数会发生变化

    ### 针对各个时间断面进行水力计算
    # 模型固定参数从设备表中读取，模型变量从时间序列表中读取
    for t in range(NumOfTime):
        # 载荷节点压力赋予初值
        p_base = CaseData['TS_NodePre'][BalanceNodeIndex[0]][t + 2]    # 此处取自时间序列表数值，与单断面计算不同
        for i in range(RowOfLoadNode):
            CaseData['TS_NodePre'][LoadNodeIndex[i]][t + 2] = p_base + 0.01 * (i+1)

        # 获得不包含平衡节点的注入列向量
        NodeInjection = np.zeros(shape=(RowOfLoadNode, 1))
        for i in range(RowOfLoadNode):
            NodeInjection[i][0] = CaseData['TS_NodeInjec'][LoadNodeIndex[i]][t + 2]

        # 计算电动阀门当前的阻抗（本程序利用【R^(2(1-x))/K^2】形式的阀门阻抗曲线）
        # 开度数据取自时间序列表，阻抗存在设备表中
        RowOfEleVal = CaseData['EleVal'].shape[0]
        for i in range(RowOfEleVal):
            EleVal_R = CaseData['EleVal'][i][2]
            EleVal_K = CaseData['EleVal'][i][3]
            EleVal_OpenDegree = CaseData['TS_EleValOpen'][i][t + 2]
            CaseData['EleVal'][i][5] = EleVal_R ** (2 * (1 - EleVal_OpenDegree)) / EleVal_K ** 2

        ########     利用牛拉法进行迭代计算      ########
        TolEro = CaseData['ErorrTor']
        #支路流量列向量
        BranchFlow = np.zeros(shape=(RowOfBranch, 1))
        #节点流量不平衡量列向量
        DeltNodeFlow = np.zeros(shape=(RowOfLoadNode, 1))

        for i in range(1000):
            print(i)
            # 计算各个支路流量
            for j in range(RowOfBranch):
                # 以下的计算假设配置泵（工频泵或变频泵）的支路不会再配置电动阀门
                if CaseData['Branch'][j][4] == 0:  # 未配置水泵的支路
                    if CaseData['TS_NodePre'][int(CaseData['Branch'][j][1])-1][t + 2] > CaseData['TS_NodePre'][int(CaseData['Branch'][j][2])-1][t + 2]:   # 注意模型变量从时间序列表中取值
                        sign = 1
                    else:
                        sign = -1

                    if CaseData['Branch'][j][8] == 0:  # 未配置电动阀门的支路
                        CaseData['TS_BranchFlow'][j][t + 2] = sign * (sign * (CaseData['TS_NodePre'][int(CaseData['Branch'][j][1])-1][t + 2]-CaseData['TS_NodePre'][int(CaseData['Branch'][j][2])-1][t + 2])/CaseData['Branch'][j][3]) ** 0.5
                    else:    # 配置电动阀门的支路
                        VauleIndex = CaseData['Branch'][j][9]  # 获取阀门编号
                        ValueResist = CaseData['EleVal'][int(VauleIndex - 1)][5]
                        CaseData['TS_BranchFlow'][j][t + 2] = sign * (sign * (CaseData['TS_NodePre'][int(CaseData['Branch'][j][1]) - 1][t + 2] - \
                                                                   CaseData['TS_NodePre'][int(CaseData['Branch'][j][2]) - 1][t + 2]) / \
                                                           (CaseData['Branch'][j][3] + ValueResist)) ** 0.5

                elif CaseData['Branch'][j][4] == 1:  # 配置工频水泵的支路
                    Pump_Index = CaseData['Branch'][j][7]
                    Pump_a0 = CaseData['ConPump'][int(Pump_Index - 1)][2]
                    Pump_a1 = CaseData['ConPump'][int(Pump_Index - 1)][3]
                    Pump_a2 = CaseData['ConPump'][int(Pump_Index - 1)][4]
                    InletPre = CaseData['TS_NodePre'][int(CaseData['Branch'][j][1]) - 1][t + 2]  # 支路进口压力
                    OutletPre = CaseData['TS_NodePre'][int(CaseData['Branch'][j][2]) - 1][t + 2]  # 支路出口压力
                    Resist = CaseData['Branch'][j][3]  # 支路阻力特性系数
                    CaseData['TS_BranchFlow'][j][t + 2] = (-1 * Pump_a1 - (Pump_a1**2 - 4 *(Pump_a2 - Resist)*(Pump_a0 + InletPre - OutletPre))**0.5)/2/(Pump_a2 - Resist)

                else:   # 配置变频水泵的支路
                    Pump_Index = CaseData['Branch'][j][7]
                    Pump_a0 = CaseData['VarPump'][int(Pump_Index - 1)][2]
                    Pump_a1 = CaseData['VarPump'][int(Pump_Index - 1)][3]
                    Pump_a2 = CaseData['VarPump'][int(Pump_Index - 1)][4]
                    FreRatio = CaseData['TS_VarPumpFre'][int(Pump_Index - 1)][t + 2] / 50 # 当前频率与额定频率比值
                    InletPre = CaseData['TS_NodePre'][int(CaseData['Branch'][j][1]) - 1][t + 2]  # 支路进口压力
                    OutletPre = CaseData['TS_NodePre'][int(CaseData['Branch'][j][2]) - 1][t + 2]  # 支路出口压力
                    Resist = CaseData['Branch'][j][3]  # 支路阻力特性系数
                    CaseData['TS_BranchFlow'][j][t + 2] = (-1 * Pump_a1 * FreRatio - (
                            (Pump_a1 * FreRatio) ** 2 - 4 * (Pump_a2 - Resist) * (Pump_a0 * FreRatio ** 2 + InletPre - OutletPre)) ** 0.5) / 2 / (
                                                           Pump_a2 - Resist)

                BranchFlow[j][0] = CaseData['TS_BranchFlow'][j][t + 2]

            # 计算节点流量不平衡量
            DeltNodeFlow = np.dot(A, BranchFlow) - NodeInjection
            AbsDeltNodeFlow = abs(DeltNodeFlow)
            MaxAbsDeltNodeFlow = np.argmax(AbsDeltNodeFlow)
            # print("节点1不平衡量", DeltNodeFlow[0, 0])
            # print("节点2不平衡量", DeltNodeFlow[1, 0])
            print(MaxAbsDeltNodeFlow)
            print(AbsDeltNodeFlow[MaxAbsDeltNodeFlow][0])

            if AbsDeltNodeFlow[MaxAbsDeltNodeFlow][0] < TolEro:
                break

            #  计算雅克比矩阵
            D = np.diag([1.0]*RowOfBranch)
            for j in range(RowOfBranch):
                D1 = CaseData['TS_NodePre'][int(CaseData['Branch'][j][1])-1][t + 2] - CaseData['TS_NodePre'][int(CaseData['Branch'][j][2])-1][t + 2]
                D[j][j] = 0.5 * CaseData['TS_BranchFlow'][j][t + 2] / D1
            Jac1 = -1 * np.dot(A, D)
            Jac = np.dot(Jac1, A.T)

            # 遍历支路表，基于配置工频泵和变频泵的支路对雅克比矩阵进行更新
            for j in range(RowOfBranch):
                if CaseData['Branch'][j][4] == 1:  # 配置工频水泵的支路
                    Pump_Index = CaseData['Branch'][j][7]
                    Pump_a0 = CaseData['ConPump'][int(Pump_Index - 1)][2]
                    Pump_a1 = CaseData['ConPump'][int(Pump_Index - 1)][3]
                    Pump_a2 = CaseData['ConPump'][int(Pump_Index - 1)][4]
                    InletNode = int(CaseData['Branch'][j][1]) # 支路进口节点编号
                    OutletNode = int(CaseData['Branch'][j][2])  # 支路出口节点编号
                    InletPre = CaseData['TS_NodePre'][InletNode - 1][t + 2]  # 支路进口压力
                    OutletPre = CaseData['TS_NodePre'][OutletNode - 1][t + 2]  # 支路出口压力
                    Resist_j = CaseData['Branch'][j][3]  # 支路阻力特性系数
                    Flow_j = CaseData['TS_BranchFlow'][j][t + 2]
                    UpDateValue = 1/2 * Flow_j / (InletPre - OutletPre) - 1 / (Pump_a1 ** 2 - 4 *(Pump_a2 - Resist_j) *(Pump_a0 + InletPre - OutletPre))**0.5

                    # 遍历LoadNodeIndex确定水泵进出口节点在非参考节点中的位置
                    Position_InletNode =  -1 # 初始化
                    Position_OutletNode = -1 # 初始化
                    for k in range(RowOfLoadNode):
                        if LoadNodeIndex[k] == (InletNode - 1):
                            Position_InletNode = k
                        if LoadNodeIndex[k] == (OutletNode - 1):
                            Position_OutletNode = k

                    if Position_InletNode != -1 and Position_OutletNode != -1:    # 支路进出口节点都不是参考节点
                        Jac[Position_InletNode][Position_InletNode] += UpDateValue
                        Jac[Position_InletNode][Position_OutletNode] -= UpDateValue
                        Jac[Position_OutletNode][Position_InletNode] -= UpDateValue
                        Jac[Position_OutletNode][Position_OutletNode] += UpDateValue

                    if Position_InletNode != -1 and Position_OutletNode == -1:
                        Jac[Position_InletNode][Position_InletNode] += UpDateValue

                    if Position_InletNode == -1 and Position_OutletNode != -1:
                        Jac[Position_OutletNode][Position_OutletNode] += UpDateValue

                if CaseData['Branch'][j][4] == 2:  # 配置变频水泵的支路
                    Pump_Index = CaseData['Branch'][j][7]
                    Pump_a0 = CaseData['VarPump'][int(Pump_Index - 1)][2]
                    Pump_a1 = CaseData['VarPump'][int(Pump_Index - 1)][3]
                    Pump_a2 = CaseData['VarPump'][int(Pump_Index - 1)][4]
                    FreRatio = CaseData['TS_VarPumpFre'][int(Pump_Index - 1)][t + 2] / 50
                    InletNode = int(CaseData['Branch'][j][1])  # 支路进口节点编号
                    OutletNode = int(CaseData['Branch'][j][2])  # 支路出口节点编号
                    InletPre = CaseData['TS_NodePre'][InletNode - 1][t + 2]  # 支路进口压力
                    OutletPre = CaseData['TS_NodePre'][OutletNode - 1][t + 2]  # 支路出口压力
                    Resist_j = CaseData['Branch'][j][3]  # 支路阻力特性系数
                    Flow_j = CaseData['TS_BranchFlow'][j][t + 2]
                    UpDateValue = 1 / 2 * Flow_j / (InletPre - OutletPre) - 1 / (
                            (Pump_a1 * FreRatio) ** 2 - 4 * (Pump_a2 - Resist_j) * (Pump_a0 * FreRatio ** 2 + InletPre - OutletPre)) ** 0.5

                    # 遍历LoadNodeIndex确定水泵进出口节点在非参考节点中的位置
                    Position_InletNode = -1  # 初始化
                    Position_OutletNode = -1  # 初始化
                    for k in range(RowOfLoadNode):
                        if LoadNodeIndex[k] == (InletNode - 1):
                            Position_InletNode = k
                        if LoadNodeIndex[k] == (OutletNode - 1):
                            Position_OutletNode = k

                    if Position_InletNode != -1 and Position_OutletNode != -1:  # 支路进出口节点都不是参考节点
                        Jac[Position_InletNode][Position_InletNode] += UpDateValue
                        Jac[Position_InletNode][Position_OutletNode] -= UpDateValue
                        Jac[Position_OutletNode][Position_InletNode] -= UpDateValue
                        Jac[Position_OutletNode][Position_OutletNode] += UpDateValue

                    if Position_InletNode != -1 and Position_OutletNode == -1:
                        Jac[Position_InletNode][Position_InletNode] += UpDateValue

                    if Position_InletNode == -1 and Position_OutletNode != -1:
                        Jac[Position_OutletNode][Position_OutletNode] += UpDateValue

            Jac = np.mat(Jac)

            #计算压力变化量
            DeltPre = Jac.I * (-1) * DeltNodeFlow


            # 更新压力值
            for j in range(RowOfLoadNode):
                CaseData['TS_NodePre'][LoadNodeIndex[j]][t + 2] = CaseData['TS_NodePre'][LoadNodeIndex[j]][t + 2] + DeltPre[j][0]
            # print(DeltPre[27, 0])
            # print(CaseData['TS_NodePre'][27][t + 2])
        # 计算参考节点注入流量
        for i in range(RowOfBalanceNode):
            for j in range(RowOfBranch):
                if CaseData['Branch'][j][1] == BalanceNodeIndex[i] + 1:
                    CaseData['TS_NodeInjec'][BalanceNodeIndex[i]][t + 2] = CaseData['TS_NodeInjec'][BalanceNodeIndex[i]][t + 2] + CaseData['TS_BranchFlow'][j][t + 2]
                if CaseData['Branch'][j][2] == BalanceNodeIndex[i] + 1:
                    CaseData['TS_NodeInjec'][BalanceNodeIndex[i]][t + 2] = CaseData['TS_NodeInjec'][BalanceNodeIndex[i]][t + 2] - CaseData['TS_BranchFlow'][j][t + 2]

        # 计算水泵流量和扬程
        RowOfConPump = CaseData['ConPump'].shape[0]
        for i in range(RowOfConPump):
            Punb_i_Branch = CaseData['ConPump'][i][1]  # 所在支路编号
            Punb_i_Flow = CaseData['TS_BranchFlow'][int(Punb_i_Branch) - 1][t + 2]
            CaseData['TS_ConPumpFlow'][i][t + 2] = Punb_i_Flow

            Pump_a0 = CaseData['ConPump'][i][2]
            Pump_a1 = CaseData['ConPump'][i][3]
            Pump_a2 = CaseData['ConPump'][i][4]
            CaseData['TS_ConPumpHead'][i][t + 1] = Pump_a0 + Pump_a1 * Punb_i_Flow + Pump_a2 * Punb_i_Flow ** 2

        RowOfVarPump = CaseData['VarPump'].shape[0]
        for i in range(RowOfVarPump):
            Punb_i_Branch = CaseData['VarPump'][i][1]  # 所在支路编号
            Punb_i_Flow = CaseData['TS_BranchFlow'][int(Punb_i_Branch) - 1][t + 2]
            CaseData['TS_VarPumpFlow'][i][t + 2] = Punb_i_Flow

            Pump_a0 = CaseData['VarPump'][i][2]
            Pump_a1 = CaseData['VarPump'][i][3]
            Pump_a2 = CaseData['VarPump'][i][4]
            FreRatio = CaseData['TS_VarPumpFre'][i][t + 2] / 50
            CaseData['TS_VarPumpHead'][i][t + 2] = Pump_a0 * FreRatio ** 2 + Pump_a1 * FreRatio * Punb_i_Flow + Pump_a2 * Punb_i_Flow ** 2


    # 模型变量时序结果写入计算结果表格
    wb = openpyxl.load_workbook(ResultFile)

    # 节点压力
    ws_1 = wb['TS_NodePre']
    RowOfTS_NodePre = CaseData['TS_NodePre'].shape[0]
    ColOfTS_NodePre = CaseData['TS_NodePre'].shape[1]
    for i in range(RowOfTS_NodePre):
        for j in range(ColOfTS_NodePre):
            ws_1.cell(2 + i, 1 + j).value = CaseData['TS_NodePre'][i][j]

    # 节点注入
    ws_2 = wb['TS_NodeInjec']
    RowOfTS_NodeInjec = CaseData['TS_NodeInjec'].shape[0]
    ColOfTS_NodeInjec = CaseData['TS_NodeInjec'].shape[1]
    for i in range(RowOfTS_NodeInjec):
        for j in range(ColOfTS_NodeInjec):
            ws_2.cell(2 + i, 1 + j).value = CaseData['TS_NodeInjec'][i][j]

    # 支路流量
    ws_3 = wb['TS_BranchFlow']
    RowOfTS_BranchFlow = CaseData['TS_BranchFlow'].shape[0]
    ColOfTS_BranchFlow = CaseData['TS_BranchFlow'].shape[1]
    for i in range(RowOfTS_BranchFlow):
        for j in range(ColOfTS_BranchFlow):
            ws_3.cell(2 + i, 1 + j).value = CaseData['TS_BranchFlow'][i][j]

    # 工频水泵流量
    ws_4 = wb['TS_ConPumpFlow']
    RowOfTS_ConPumpFlow = CaseData['TS_ConPumpFlow'].shape[0]
    ColOfTS_ConPumpFlow = CaseData['TS_ConPumpFlow'].shape[1]
    for i in range(RowOfTS_ConPumpFlow):
        for j in range(ColOfTS_ConPumpFlow):
            ws_4.cell(2 + i, 1 + j).value = CaseData['TS_ConPumpFlow'][i][j]

    # 工频水泵扬程
    ws_5 = wb['TS_ConPumpHead']
    RowOfTS_ConPumpHead = CaseData['TS_ConPumpHead'].shape[0]
    ColOfTS_ConPumpHead = CaseData['TS_ConPumpHead'].shape[1]
    for i in range(RowOfTS_ConPumpHead):
        for j in range(ColOfTS_ConPumpHead):
            ws_5.cell(2 + i, 1 + j).value = CaseData['TS_ConPumpHead'][i][j]

    # 变频水泵流量
    ws_6 = wb['TS_VarPumpFlow']
    RowOfTS_VarPumpFlow = CaseData['TS_VarPumpFlow'].shape[0]
    ColOfTS_VarPumpFlow = CaseData['TS_VarPumpFlow'].shape[1]
    for i in range(RowOfTS_VarPumpFlow):
        for j in range(ColOfTS_VarPumpFlow):
            ws_6.cell(2 + i, 1 + j).value = CaseData['TS_VarPumpFlow'][i][j]

    # 变频水泵扬程
    ws_7 = wb['TS_VarPumpHead']
    RowOfTS_VarPumpHead = CaseData['TS_VarPumpHead'].shape[0]
    ColOfTS_VarPumpHead = CaseData['TS_VarPumpHead'].shape[1]
    for i in range(RowOfTS_VarPumpHead):
        for j in range(ColOfTS_VarPumpHead):
            ws_7.cell(2 + i, 1 + j).value = CaseData['TS_VarPumpHead'][i][j]

    wb.save(ResultFile)

    return CaseData

if __name__ == '__main__':
    # # 69节点热网水力计算【2024-3-3】
    # Network = IHES_ReadNetworkPara_V2.ReadPara('Chapter 1_69节点热力系统算例-拓扑数据(删除69-31支路和100-138支路).xlsx', 'Chapter 1_69节点热力系统算例-12-11.45-时序数据（用于水力计算)(删除69-31支路和100-138支路).xlsx')
    # Result = Hydraulic_Calculation(Network, 'Chapter 1_69节点热力系统算例-12-11.45-时序数据（水力计算结果) .xlsx', 1)

    # 热网水力计算【2024 - 3 - 3】
    Network = IHES_ReadNetworkPara_V2.ReadPara('Chapter 1_69节点热力系统算例-拓扑数据.xlsx',
                                               'Chapter 1_69节点热力系统算例-12-11.45-时序数据（用于水力计算).xlsx')
    Result = Hydraulic_Calculation(Network, 'Chapter 1_69节点热力系统算例-12-11.45-时序数据（水力计算结果) .xlsx', 1)

    # # 热网水力计算【2024 - 3 - 5】
    # # 将时序表中的CHP处变频泵由50修改为40后进行计算
    # Network = IHES_ReadNetworkPara_V2.ReadPara('Chapter 1_69节点热力系统算例-拓扑数据.xlsx',
    #                                            'Chapter 1_69节点热力系统算例-12-11.45-时序数据（用于水力计算).xlsx')
    # Network['TS_VarPumpFre'][0, 2] = 45
    # Result = Hydraulic_Calculation(Network, 'Chapter 1_69节点热力系统算例-12-11.45-时序数据（水力计算结果)-CHP机组水泵频率45，其余50.xlsx', 1)






