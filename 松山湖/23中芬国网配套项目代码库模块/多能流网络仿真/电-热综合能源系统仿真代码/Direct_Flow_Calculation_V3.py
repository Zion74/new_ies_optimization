# 本程序用于计算计及节点相角和支路功率的直流潮流方程【2023-12】
# 直流潮流紧凑形式为AX=b。X中的未知变量包括参考节点处平衡机组出力、非参考节点相角和支路功率
# 考虑了风电机组表【2023-12-15】
# 考虑了CHP机组单独建表，目前的平衡节点只有火电机组担任【2023-12-25】
# 增加针对直流潮流模型的灵敏度计算代码【2024-04-26】
# 增加针对多个断面进行直流潮流计算的代码【2024-04-29】

import numpy as np
import IHES_ReadNetworkPara_V2
from openpyxl import load_workbook
import gurobipy as gp
from gurobipy import GRB

def DirectFlow(CaseData):
    NumofNode = CaseData['EleNode'].shape[0]
    NumofBranch = CaseData['EleBranch'].shape[0]

    A = np.zeros(shape=(NumofNode + NumofBranch, NumofNode + NumofBranch))
    b = np.zeros(shape=(NumofNode + NumofBranch, 1))

    # 遍历各个节点，利用节点功率方程对矩阵A和b中的元素赋值
    for i in range(NumofNode):
        if CaseData['EleNode'][i, 1] == 1: # 参考节点，电压幅值已知、功率未知
            A[i, i] = 1

        for j in range(NumofBranch): # 遍历支路
            if CaseData['EleBranch'][j, 1] == i + 1:
                A[i, NumofNode + j] = -1
            if CaseData['EleBranch'][j, 2] == i + 1:
                A[i, NumofNode + j] = 1

        for j in range(CaseData['EleLoad'].shape[0]): # 遍历负荷表
            if CaseData['EleLoad'][j, 1] == i + 1:
                b[i] += CaseData['EleLoad'][j, 4]

        for j in range(CaseData['EleGen'].shape[0]):    # 遍历发电机表（不包含CHP机组）
            if CaseData['EleGen'][j, 1] == i + 1 and CaseData['EleGen'][j, 15] == 0:
                b[i, 0] -= CaseData['EleGen'][j, 16]

        for j in range(CaseData['EleCHP'].shape[0]):    # 遍历CHP表
            if CaseData['EleCHP'][j, 1] == i + 1:
                b[i, 0] -= CaseData['EleCHP'][j, 14]

        for j in range(CaseData['EleWindUnit'].shape[0]):    # 遍历风电机表
            if CaseData['EleWindUnit'][j, 1] == i + 1:
                b[i, 0] -= CaseData['EleWindUnit'][j, 4]

    # 遍历各个支路，利用支路功率特性方程对矩阵A和b中元素赋值
    for i in range(NumofBranch):
        A[NumofNode + i, NumofNode + i] = -1 / CaseData['EleBase']  # 支路阻抗参数对应的是支路功率标幺值

        Inlet_i = int(CaseData['EleBranch'][i, 1])
        Outlet_i = int(CaseData['EleBranch'][i, 2])

        if CaseData['EleNode'][Inlet_i - 1, 1] == 0: # 进口节点非参考节点
            A[NumofNode + i, Inlet_i - 1] = 1 / CaseData['EleBranch'][i, 3]

        else:
            b[NumofNode + i, 0] += -1 / CaseData['EleBranch'][i, 3] * CaseData['EleNode'][Inlet_i - 1, 5]

        if CaseData['EleNode'][Outlet_i - 1, 1] == 0:  # 出口节点非参考节点
            A[NumofNode + i, Outlet_i - 1] = -1 / CaseData['EleBranch'][i, 3]

        else:
            b[NumofNode + i, 0] += 1 / CaseData['EleBranch'][i, 3] * CaseData['EleNode'][Inlet_i - 1, 5]

    A = np.mat(A)
    Result = np.dot(A.I, b)
    print(Result)

    # Result中的结果写回CaseData中
    # 节点处的未知变量
    for i in range(NumofNode):
        if CaseData['EleNode'][i, 1] == 1: # 参考节点,需要将平衡节点电出力赋值到平衡机组
            for j in range(CaseData['EleGen'].shape[0]):
                if CaseData['EleGen'][j, 1] == i + 1:
                    CaseData['EleGen'][j, 16] = Result[i]
        else:
            CaseData['EleNode'][i, 5] = Result[i]

    # 支路上的未知变量
    for i in range(NumofBranch):
        CaseData['EleBranch'][i, 6] = Result[i + NumofNode]

    return CaseData

def DirectFlowSen(CaseData, ResutlFile):    # 计算直流潮流的节点净注入功率导致的灵敏度矩阵
    # CaseData: 电力系统拓扑数据
    # ResutlFile：存储灵敏度结果的表格

    NumofNode = CaseData['EleNode'].shape[0]
    NumofBranch = CaseData['EleBranch'].shape[0]

    A = np.zeros(shape=(NumofNode + NumofBranch, NumofNode + NumofBranch))
    C = np.zeros(shape=(NumofNode + NumofBranch, NumofNode - 1))

    # 遍历各个节点，利用节点功率方程对矩阵A赋值
    for i in range(NumofNode):
        if CaseData['EleNode'][i, 1] == 1:  # 参考节点，电压幅值已知、功率未知
            A[i, i] = 1

        for j in range(NumofBranch):  # 遍历支路
            if CaseData['EleBranch'][j, 1] == i + 1:
                A[i, NumofNode + j] = -1
            if CaseData['EleBranch'][j, 2] == i + 1:
                A[i, NumofNode + j] = 1

    # 遍历各个支路，利用支路功率特性方程对矩阵A
    for i in range(NumofBranch):
        A[NumofNode + i, NumofNode + i] = -1 / CaseData['EleBase']  # 支路阻抗参数对应的是支路功率标幺值

        Inlet_i = int(CaseData['EleBranch'][i, 1])
        Outlet_i = int(CaseData['EleBranch'][i, 2])

        if CaseData['EleNode'][Inlet_i - 1, 1] == 0:  # 进口节点非参考节点
            A[NumofNode + i, Inlet_i - 1] = 1 / CaseData['EleBranch'][i, 3]

        if CaseData['EleNode'][Outlet_i - 1, 1] == 0:  # 出口节点非参考节点
            A[NumofNode + i, Outlet_i - 1] = -1 / CaseData['EleBranch'][i, 3]

    #  对矩阵C赋值
    RefNodeLable = 0
    for i in range(NumofNode):
        if CaseData['EleNode'][i, 1] == 0:  # 非参考节点
            C[i, i - RefNodeLable] = -1
        else:
            RefNodeLable += 1

    A = np.mat(A)
    Sensitivity = np.dot(A.I, C)

    wb = load_workbook(ResutlFile)
    ws_1 = wb['Sensitivity']
    for i in range(Sensitivity.shape[0]):
        for j in range(Sensitivity.shape[1]):
            ws_1.cell(i + 2, j + 2).value = Sensitivity[i, j]

    wb.save(ResutlFile)

    return Sensitivity

def DirectFlowVI(CaseData, Sensitivity, ResultFile):
    ########     支路功率脆弱性指标计算
    # 节点注入向上调，支路功率脆弱性指标
    BranchPowerVI_Up = np.zeros(shape=(CaseData['EleBranch'].shape[0], CaseData['EleNode'].shape[0] - 1))
    for i in range(BranchPowerVI_Up.shape[0]):
        for j in range(BranchPowerVI_Up.shape[1]):
            if Sensitivity[i + CaseData['EleNode'].shape[0], j] >= 0:
                BranchPowerVI_Up[i, j] = Sensitivity[i + CaseData['EleNode'].shape[0], j] / (
                            CaseData['EleBranch'][i, 4] - CaseData['EleBranch'][i, 6])

            else:
                BranchPowerVI_Up[i, j] = Sensitivity[i + CaseData['EleNode'].shape[0], j] / (
                            CaseData['EleBranch'][i, 6] - CaseData['EleBranch'][i, 5])

    # 节点注入向下调，支路功率脆弱性指标
    BranchPowerVI_Dn = np.zeros(shape=(CaseData['EleBranch'].shape[0], CaseData['EleNode'].shape[0] - 1))
    for i in range(BranchPowerVI_Dn.shape[0]):
        for j in range(BranchPowerVI_Dn.shape[1]):
            if Sensitivity[i + CaseData['EleNode'].shape[0], j] >= 0:
                BranchPowerVI_Dn[i, j] = Sensitivity[i + CaseData['EleNode'].shape[0], j] / (
                            CaseData['EleBranch'][i, 5] - CaseData['EleBranch'][i, 6])

            else:
                BranchPowerVI_Dn[i, j] = Sensitivity[i + CaseData['EleNode'].shape[0], j] / (
                            CaseData['EleBranch'][i, 6] - CaseData['EleBranch'][i, 4])


    wb = load_workbook(ResultFile)
    ws_1 = wb['BranchFlowVI']
    for i in range(BranchPowerVI_Up.shape[0]):
        for j in range(BranchPowerVI_Up.shape[1]):
            ws_1.cell(i + 3, j + 2).value = BranchPowerVI_Up[i, j]
            ws_1.cell(i + 15, j + 2).value = BranchPowerVI_Dn[i, j]


    wb.save(ResultFile)

def EleSecControl(CaseData, Sensitivity, BranchPower_margin, ResultFile):
    # Sensitivity:节点变量和支路变量灵敏度，应用支路功率灵敏度度是注意行号
    # BranchPower_margin:支路功率安全裕度

    # 计算状态变量预期变化量和权重
    NumOfBranch = CaseData['EleBranch'].shape[0]
    StateVarAdj = []    # 状态变量预期变化量
    StateVarAdjWeight = []  # 状态变量预期变化量权重

    for i in range(NumOfBranch):
        BranchPower_i = CaseData['EleBranch'][i, 6]
        BranchPower_i_max = CaseData['EleBranch'][i, 4]
        BranchPower_i_min = CaseData['EleBranch'][i, 5]

        if BranchPower_i >= BranchPower_i_min + BranchPower_margin and BranchPower_i <= BranchPower_i_max - BranchPower_margin:
            AdjValue = 0
            AdjWeight = 0.1

        elif BranchPower_i >= BranchPower_i_min and BranchPower_i <= BranchPower_i_min + BranchPower_margin:
            AdjValue = BranchPower_i_min + BranchPower_margin - BranchPower_i
            AdjWeight = 0.5

        elif BranchPower_i >= BranchPower_i_max - BranchPower_margin and BranchPower_i <= BranchPower_i_max:
            AdjValue = BranchPower_i_max - BranchPower_margin - BranchPower_i
            AdjWeight = 0.5

        elif BranchPower_i < BranchPower_i_min:
            AdjValue = BranchPower_i_min + BranchPower_margin - BranchPower_i
            AdjWeight = 10

        else:
            AdjValue = BranchPower_i_max - BranchPower_margin - BranchPower_i
            AdjWeight = 10

        StateVarAdj.append(AdjValue)
        StateVarAdjWeight.append(AdjWeight)

    DiagWeight = np.eye(NumOfBranch)
    row, col = np.diag_indices_from(DiagWeight)
    DiagWeight[row, col] = StateVarAdjWeight

    # 控制变量（节点注入）和状态变量（支路功率）当前值,控制变量（节点净注入功率）范围计算
    # 注意控制变量的排序需要和灵敏度中控制变量排序相同
    ContrVar = []   # 节点净注入功率当前值
    NumOfNonRef = 0  # 非参考节点数目
    ContrVar_Max = []  # 节点注入最大值：节点所连发电机上限之和
    ContrVar_Min = []   # 节点注入最小值：节点所连发电机下限-所连负荷功率

    for i in range(CaseData['EleNode'].shape[0]):
        NetInPower = 0
        NetInPower_Max = 0
        NetInPower_Min = 0
        if CaseData['EleNode'][i, 1] == 0:  # 非参考节点
            NumOfNonRef += 1
            for j in range(CaseData['EleLoad'].shape[0]):   # 遍历负荷表
                if CaseData['EleLoad'][j, 1] == i + 1:
                    NetInPower -= CaseData['EleLoad'][j, 4]
                    NetInPower_Min -= CaseData['EleLoad'][j, 4]

            for j in range(CaseData['EleCHP'].shape[0]):    # 遍历CHP机组表
                if CaseData['EleCHP'][j, 1] == i + 1:
                    NetInPower += CaseData['EleCHP'][j, 14]
                    NetInPower_Max += CaseData['EleCHP'][j, 4]
                    NetInPower_Min += CaseData['EleCHP'][j, 5]

            for j in range(CaseData['EleWindUnit'].shape[0]):   # 遍历风电机组表
                if CaseData['EleWindUnit'][j, 1] == i + 1:
                    NetInPower += CaseData['EleWindUnit'][j, 4]
                    NetInPower_Max += CaseData['EleWindUnit'][j, 2]
                    NetInPower_Min += CaseData['EleWindUnit'][j, 5]

            for j in range(CaseData['EleGen'].shape[0]):    # 遍历火电机组表
                if CaseData['EleGen'][j, 1] == i + 1:
                    NetInPower += CaseData['EleGen'][j, 16]
                    NetInPower_Max += CaseData['EleGen'][j, 4]
                    NetInPower_Min += CaseData['EleGen'][j, 5]

            ContrVar.append(NetInPower)
            ContrVar_Max.append(NetInPower_Max)
            ContrVar_Min.append(NetInPower_Min)

    StateVar = []   # 支路功率当前值
    for i in range(NumOfBranch):
        StateVar.append(CaseData['EleBranch'][i, 6])


    # 建立优化模型，求取安全校正方案。优化直流潮流模型灵敏度度固定，所以单次求解优化即可求得安全校正方案
    m = gp.Model('Security control by optimization model')

    # 1、建立决策变量
    I = list(range(NumOfNonRef))    # 非参考节点数目
    DeltContrVar_Cal = m.addVars(I, lb=-GRB.INFINITY, vtype=GRB.CONTINUOUS,name='Calculated change of control variable')  # 添加lb = -GRB.INFINITY能够使功率取到负值

    # 2、建立约束条件
    m.addConstrs(DeltContrVar_Cal[i] + ContrVar[i] <= ContrVar_Max[i] for i in I)
    m.addConstrs(DeltContrVar_Cal[i] + ContrVar[i] >= ContrVar_Min[i] for i in I)

    # 3、建立目标函数
    J = list(range(NumOfBranch))
    m.setObjective(sum((StateVarAdj[j] - sum(Sensitivity[j + CaseData['EleNode'].shape[0], i] * DeltContrVar_Cal[i] for i in I)) ** 2
                       * DiagWeight[j, j] for j in J), GRB.MINIMIZE)
    m.Params.MIPGap = 0.000001
    m.optimize()

    ControResult = []
    for i in range(NumOfNonRef):
        ControResult.append(DeltContrVar_Cal[i].X)

    wb = load_workbook(ResultFile)
    ws = wb['ControResult']
    for i in range(NumOfNonRef):
        ws.cell(2, 2 + i).value = ControResult[i]
    wb.save(ResultFile)

    return ControResult

def DirectFlowMultiTime(CaseData, NumOfTime, ResultFile):
    # CaseData 包含拓扑和时序数据
    # NumOfTime：需要进行计算的时间断面数量
    # ResultFile：存储多时段计算结果的文件

    for t in range(NumOfTime):
        # 将时序表中的电负荷、风电、CHP和非平衡机组功率赋值到拓扑表
        for i in range(CaseData['EleLoad'].shape[0]):   # 遍历电负荷表
            CaseData['EleLoad'][i, 4] = CaseData['TS_EleLoad'][i, t + 2]

        for i in range(CaseData['EleGen'].shape[0]):    # 遍历发电机表（不包含CHP机组）
            if CaseData['EleGen'][i, 15] == 0:  # 非平衡火电机组
                CaseData['EleGen'][i, 16] = CaseData['TS_EleGen'][i, t + 2]

        for i in range(CaseData['EleCHP'].shape[0]):    # 遍历CHP表
            CaseData['EleCHP'][i, 14] = CaseData['TS_EleCHP'][i, t + 2]

        for i in range(CaseData['EleWindUnit'].shape[0]):    # 遍历风电机表
            CaseData['EleWindUnit'][i, 4] = CaseData['TS_Wind'][i, t + 2]

        for i in range(CaseData['EleNode'].shape[0]):   # 平衡节点的相角赋值到拓扑表
            if CaseData['EleNode'][i, 1] == 1:  # 参考节点
                CaseData['EleNode'][i, 5] = CaseData['TS_EleNode'][i, t + 2]

        Result = DirectFlow(CaseData)

        #### 计算所得节点相角、支路功率和平衡计算功率写回时序表
        for i in range(CaseData['EleNode'].shape[0]):
            if CaseData['EleNode'][i, 1] == 1:  # 参考节点,需要将平衡节点电出力写入时序表
                for j in range(CaseData['EleGen'].shape[0]):
                    if CaseData['EleGen'][j, 1] == i + 1:
                        CaseData['TS_EleGen'][j, t + 2] = Result['EleGen'][j, 16]
            else:
                CaseData['TS_EleNode'][i, t + 2] = Result['EleNode'][i, 5]

        # 支路上的未知变量
        for i in range(CaseData['EleBranch'].shape[0]):
            CaseData['TS_EleBranch'][i, t + 2] = Result['EleBranch'][i, 6]

    # 时序结果写回表格
    wb = load_workbook(ResultFile)
    ws_1 = wb['TS_EleGen']  # 发电机结果
    for i in range(CaseData['EleGen'].shape[0]):
        ws_1.cell(2 + i, 1).value = i + 1
        for t in range(NumOfTime):
            ws_1.cell(2 + i, 3 + t).value = CaseData['TS_EleGen'][i, t + 2]

    ws_2 = wb['TS_EleNode']     # 节点相角写回表格
    for i in range(CaseData['EleNode'].shape[0]):
        ws_2.cell(2 + i, 1).value = i + 1
        for t in range(NumOfTime):
            ws_2.cell(2 + i, 3 + t).value = CaseData['TS_EleNode'][i, t + 2]

    ws_3 = wb['TS_EleBranch']
    for i in range(CaseData['EleBranch'].shape[0]):
        ws_3.cell(2 + i, 1).value = i + 1
        for t in range(NumOfTime):
            ws_3.cell(2 + i, 3 + t).value = CaseData['TS_EleBranch'][i, t + 2]

    wb.save(ResultFile)



if __name__ == '__main__':
    # # 利用电力系统直流模型进行计算
    # CaseData = IHES_ReadNetworkPara_V2.ReadPara('6节点电力系统-14节点热力系统算例-拓扑数据.xlsx', '6节点电力系统-14节点热力系统算例-优化计算-时序数据（包含水力结果）.xlsx')
    #
    # # 电负荷功率
    # for i in range(CaseData['EleLoad'].shape[0]):
    #     CaseData['EleLoad'][i, 4] = CaseData['TS_EleLoad'][i, 2]
    # # 参考节点相角
    # CaseData['EleNode'][0, 5] = 0
    # # 工况1：各个非平衡节点发电机功率
    # CaseData['EleGen'][1, 16] = 200
    # CaseData['EleWindUnit'][0, 4] = 300
    # CaseData['EleCHP'][0, 14] = 300

    # # 工况2：各个非平衡节点发电机功率
    # CaseData['EleGen'][1, 16] = 200
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 200

    # # 工况3：电负荷与发电机功率
    # CaseData['EleLoad'][0, 4] = 400
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 200
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300


    # # 工况4：在工况3基础上，节点2处火电增加650，查看越限状态变量
    # CaseData['EleLoad'][0, 4] = 400
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 200 + 650
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300

    # # 工况5：在工况3基础上，节点5处电负荷增加150，查看越限状态变量
    # CaseData['EleLoad'][0, 4] = 400
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 + 150
    # CaseData['EleGen'][1, 16] = 200
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300

    # # # 工况6：电负荷与发电机功率
    # CaseData['EleLoad'][0, 4] = 400
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 200 + 50     #   与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300

    # # # 工况7：电负荷与发电机功率
    # CaseData['EleLoad'][0, 4] = 400 + 20   # 与工况6不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300

    # # 工况8：电负荷与发电机功率
    # CaseData['EleLoad'][0, 4] = 400 + 20   # 与工况6不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300

    # # 工况9：电负荷与发电机功率
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10   # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300
    #
    #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 12).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标.xlsx')


    # # 灵敏度计算（与工况状态无关）
    # SensitivityResult = DirectFlowSen(Result, '6节点电力系统灵敏度结果与脆弱性指标.xlsx')
    #
    # # 基于【工况3】设置的参数，计算脆弱性指标
    # DirectFlowVI(Result, SensitivityResult, '6节点电力系统灵敏度结果与脆弱性指标.xlsx')


    # ############    以工况3为基态，逐步进行安全校正
    # CaseData = IHES_ReadNetworkPara_V2.ReadPara('6节点电力系统-14节点热力系统算例-拓扑数据.xlsx',
    #                                             '6节点电力系统-14节点热力系统算例-优化计算-时序数据（包含水力结果）.xlsx')
    # # 参考节点相角
    # CaseData['EleNode'][0, 5] = 0
    # # 第1步安全校正：在工况3基础上，节点4注入变化量为-60.06794005
    # CaseData['EleLoad'][0, 4] = 400
    # CaseData['EleLoad'][1, 4] = 300 - 60.06794005
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 200
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300
    #
    #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 6).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标.xlsx')


    # ############    以工况7为基态，逐步进行安全校正
    # CaseData = IHES_ReadNetworkPara_V2.ReadPara('6节点电力系统-14节点热力系统算例-拓扑数据.xlsx',
    #                                             '6节点电力系统-14节点热力系统算例-优化计算-时序数据（包含水力结果）.xlsx')
    # # 参考节点相角
    # CaseData['EleNode'][0, 5] = 0

    # # 第1步安全校正：在工况7基础上，节点3注入变化量为55.419
    # CaseData['EleLoad'][0, 4] = 400 + 20 - 55.419
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300

    # # 第2步安全校正：在第1步校正基础上，节点5注入变化量为109.83
    # CaseData['EleLoad'][0, 4] = 400 + 20 - 55.419
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 - 109.83
    # CaseData['EleGen'][1, 16] = 250 + 50
    # CaseData['EleWindUnit'][0, 4] = 100
    # CaseData['EleCHP'][0, 14] = 300
    #
    #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 10).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标.xlsx')


    # ############    以工况9为基态，逐步进行安全校正
    # CaseData = IHES_ReadNetworkPara_V2.ReadPara('6节点电力系统-14节点热力系统算例-拓扑数据.xlsx',
    #                                             '6节点电力系统-14节点热力系统算例-优化计算-时序数据（包含水力结果）.xlsx')
    # # 参考节点相角
    # CaseData['EleNode'][0, 5] = 0

    # # 基于工况9，第一步安全校正：节点3注入增加60.00006
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 60.00006  # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300

    # # 基于工况9，第一步安全校正：节点3注入增加60.00006
    # # 基于工况9，第二步安全校正：节点5注入增加155.04
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 60.00006  # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 - 155.04
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300

    # # 基于工况9，第一步安全校正：节点3注入增加60.00006
    # # 基于工况9，第二步安全校正：节点5注入增加155.04
    # # 基于工况9，第三步安全校正：节点6注入降低56.811
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 60.00006  # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 - 155.04
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300 - 56.811
    #
    #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 15).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标.xlsx')


    # #####################           基于工况9注入参数设置，重新进行灵敏度计算、脆弱性指标计算和脆弱性指标识别脆弱状态变量、安全校正    ####################
    # CaseData = IHES_ReadNetworkPara_V2.ReadPara('6节点电力系统-14节点热力系统算例-拓扑数据.xlsx',
    #                                             '6节点电力系统-14节点热力系统算例-优化计算-时序数据（包含水力结果）.xlsx')
    # # 参考节点相角
    # CaseData['EleNode'][0, 5] = 0
    #
    # # 工况9注入参数设置
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10   # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300
    #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 2).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    #
    #
    # # 灵敏度计算（与工况状态无关）
    # SensitivityResult = DirectFlowSen(Result, '6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    #
    # # 基于【工况9】设置的参数，计算脆弱性指标
    # DirectFlowVI(Result, SensitivityResult, '6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')

    #######   脆弱性结果验证
    # ###  验证 VI_Up
    # # 改变3次节点5注入：增加350/300/250/200
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10   # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 - 200
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300

    # ###  验证 VI_Dn
    # # 改变3次节点5注入：减少250/200/150/100
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10   # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 + 100
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300
    #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 12).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')


    #######  逐步进行安全校正
    # # 基于工况9，第一步安全校正：节点3注入增加60.00006
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 60.00006  # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300

    # # 基于工况9，第一步安全校正：节点3注入增加60.00006
    # # 基于工况9，第二步安全校正：节点5注入增加155.04
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 60.00006  # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 - 155.04
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300

    # 基于工况9，第一步安全校正：节点3注入增加60.00006
    # 基于工况9，第二步安全校正：节点5注入增加155.04
    # 基于工况9，第三步安全校正：节点6注入降低56.811
    # CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 60.00006  # 与工况6不同的地方、与工况8不同的地方
    # CaseData['EleLoad'][1, 4] = 300
    # CaseData['EleLoad'][2, 4] = 300 - 155.04
    # CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    # CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    # CaseData['EleCHP'][0, 14] = 300 - 56.811
    # #
    # #
    # #  支路潮流计算
    # Result = DirectFlow(CaseData)
    # wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    # ws = wb['CalResult']
    # for i in range(Result['EleBranch'].shape[0]):
    #     ws.cell(3 + i, 16).value = Result['EleBranch'][i, 6]
    # wb.save('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')


    ########    利用优化模型计算安全校正程序
    CaseData = IHES_ReadNetworkPara_V2.ReadPara('6节点电力系统-14节点热力系统算例-拓扑数据.xlsx',
                                                '6节点电力系统-14节点热力系统算例-优化计算-时序数据（包含水力结果）.xlsx')
    # 参考节点相角
    CaseData['EleNode'][0, 5] = 0

    # 工况9注入参数设置
    CaseData['EleLoad'][0, 4] = 400 + 20 + 10   # 与工况6不同的地方、与工况8不同的地方
    CaseData['EleLoad'][1, 4] = 300
    CaseData['EleLoad'][2, 4] = 300
    CaseData['EleGen'][1, 16] = 250 + 50    #  与工况3不同的地方
    CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    CaseData['EleCHP'][0, 14] = 300

    # 潮流计算
    Result = DirectFlow(CaseData)
    # 计算灵敏度
    SensitivityResult = DirectFlowSen(Result, '6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')

    # 调用安全校正优化程序
    # 调整支路功率上限和下限值
    for i in range(Result['EleBranch'].shape[0]):
        Result['EleBranch'][i, 4] = 200
        Result['EleBranch'][i, 5] = -200

    ControResult = EleSecControl(Result, SensitivityResult, 10, '6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    print(ControResult)

    #### 验证所得安全校正策略
    CaseData['EleLoad'][0, 4] = 400 + 20 + 10 - 50.82   # 与工况6不同的地方、与工况8不同的地方
    CaseData['EleLoad'][1, 4] = 300 - 34.23
    CaseData['EleLoad'][2, 4] = 300 - 44.24
    CaseData['EleGen'][1, 16] = 250 + 50 - 92.57    #  与工况3不同的地方
    CaseData['EleWindUnit'][0, 4] = 100 + 20    # 与工况7不同的地方
    CaseData['EleCHP'][0, 14] = 300 - 40.56

    #  支路潮流计算
    Result = DirectFlow(CaseData)
    wb = load_workbook('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')
    ws = wb['CalResult']
    for i in range(Result['EleBranch'].shape[0]):
        ws.cell(3 + i, 18).value = Result['EleBranch'][i, 6]
    wb.save('6节点电力系统灵敏度结果与脆弱性指标-基于工况9参数作为基态.xlsx')



















